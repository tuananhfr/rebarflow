"""Báo cáo tính thép dạng xlsx — tái hiện trang in $B$1:$U$(20+n) của sheet D1.

Bản in tự giải thích: ghi rõ vật liệu, lớp bảo vệ, CHẾ ĐỘ TÍNH và phiên bản
tool — để ai cầm bản in cũng biết số được tính theo quy ước nào.
"""

from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from rebarflow.__version__ import __version__
from rebarflow.constants import CONCRETE, STEEL
from rebarflow.core.models import CalcMode, MaterialParams, StripDesign

_FONT = "Times New Roman"
_FILL_OK = PatternFill("solid", start_color="C6EFCE")      # xanh: đạt
_FILL_FAIL = PatternFill("solid", start_color="FFC7CE")    # đỏ: thiếu thép
_FILL_WARN = PatternFill("solid", start_color="FFEB9C")    # cam: "CT"
_THIN = Side(style="thin")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_MODE_LABEL = {
    CalcMode.EXCEL_COMPAT: "Giống Excel gốc (mặc định)",
    CalcMode.TCVN_STRICT: "TCVN chuẩn sách (tùy chọn)",
}

_HEADERS = [
    ("STT", 5), ("Tên Đài", 10), ("h đài\n(m)", 7), ("Strip", 8), ("Rộng\n(m)", 7),
    ("Tổ hợp", 8), ("Vị trí\n(m)", 7), ("M+\n(KNm)", 9), ("Tổ hợp", 8), ("Vị trí\n(m)", 7),
    ("M-\n(KNm)", 9), ("Shear\n(KN)", 9), ("Astop\n(cm2)", 8), ("Asbot\n(cm2)", 8),
    ("Ø trên\n(mm)", 7), ("a trên\n(mm)", 7), ("Ø dưới\n(mm)", 7), ("a dưới\n(mm)", 7),
    ("Check\ntrên", 7), ("Check\ndưới", 7),
]
_FIRST_COL = 2  # bảng bắt đầu từ cột B như trang in gốc


def export_report(
    designs: list[StripDesign],
    mat: MaterialParams,
    path: str,
    source: str = "",
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Thep dai"

    n_cols = len(_HEADERS)
    last_col = get_column_letter(_FIRST_COL + n_cols - 1)

    def cell(row, col, value, *, bold=False, size=11, align="center", border=False):
        c = ws.cell(row=row, column=col, value=value)
        c.font = Font(name=_FONT, size=size, bold=bold)
        c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
        if border:
            c.border = _BORDER
        return c

    # ---- tiêu đề + thông tin chung ----
    ws.merge_cells(f"B1:{last_col}1")
    cell(1, _FIRST_COL, "TÍNH TOÁN THÉP ĐÀI CỌC", bold=True, size=14)
    ws.merge_cells(f"B2:{last_col}2")
    cell(2, _FIRST_COL,
         "Tính toán sàn theo dải Strip trong SAFE theo cấu kiện chịu uốn - TCVN 5574:2012",
         size=10)
    rb = CONCRETE[mat.concrete][0]
    rs = STEEL[mat.steel][0] / 10
    info = [
        f"Nguồn: {source}" if source else "",
        f"Bê tông {mat.concrete} (Rb tra bảng = {rb:g} kG/cm²)   —   "
        f"Thép {mat.steel} (Rs = {rs:g} MPa)",
        f"Lớp bảo vệ trên {mat.cover_top_mm:g} mm / dưới {mat.cover_bot_mm:g} mm   —   "
        f"Thép sàn hầm {mat.as_ham_cm2:g} cm²   —   Chế độ tính: {_MODE_LABEL[mat.mode]}",
    ]
    r = 3
    for line in info:
        if line:
            ws.merge_cells(f"B{r}:{last_col}{r}")
            cell(r, _FIRST_COL, line, size=10, align="left")
            r += 1

    # ---- header bảng ----
    head_row = r + 1
    for i, (title, width) in enumerate(_HEADERS):
        col = _FIRST_COL + i
        cell(head_row, col, title, bold=True, size=10, border=True)
        ws.column_dimensions[get_column_letter(col)].width = width

    # ---- dữ liệu ----
    for i, d in enumerate(designs, start=1):
        row = head_row + i
        values = [
            i, d.pile_cap_name, d.h, d.env.strip, d.env.width,
            d.env.m_pos_combo, d.env.m_pos_station, d.env.m_pos,
            d.env.m_neg_combo, d.env.m_neg_station, d.env.m_neg, d.env.v_max,
            _num(d.as_top_req), _num(d.as_bot_req),
            d.dia_top, d.spacing_top, d.dia_bot, d.spacing_bot,
            _num(d.check_top), _num(d.check_bot),
        ]
        for j, v in enumerate(values):
            c = cell(row, _FIRST_COL + j, v, size=10, border=True)
            if isinstance(v, float):
                c.number_format = "0.00"
        _fill_check(ws.cell(row=row, column=_FIRST_COL + 18), d.check_top)
        _fill_check(ws.cell(row=row, column=_FIRST_COL + 19), d.check_bot)

    # ---- chân trang ----
    foot = head_row + len(designs) + 2
    ws.merge_cells(f"B{foot}:{last_col}{foot}")
    cell(foot, _FIRST_COL,
         f"Xuất từ rebarFlow v{__version__} — {datetime.now():%d/%m/%Y %H:%M}",
         size=9, align="right")

    ws.print_area = f"B1:{last_col}{foot}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    wb.save(path)


def _num(v: float | str | None) -> float | str:
    return "-" if v is None else v


def _fill_check(c, value) -> None:
    if value == "CT":
        c.fill = _FILL_WARN
    elif isinstance(value, float):
        c.fill = _FILL_OK if value >= 1 else _FILL_FAIL
