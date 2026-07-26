"""Test báo cáo xlsx: xuất từ dữ liệu golden rồi đọc lại, kiểm tra layout + số liệu."""

import pytest
from openpyxl import load_workbook

from rebarflow.core.models import CalcMode, MaterialParams, StripDesign, StripEnvelope
from rebarflow.core.rebar_calc import calc_strip
from rebarflow.export.report_xlsx import export_report


@pytest.fixture()
def report(tmp_path, golden_d1):
    designs = []
    for row in golden_d1["rows"]:
        env = StripEnvelope(
            strip=row["strip"], width=row["width"],
            m_pos=float(row["m_pos"]), m_pos_combo=row["m_pos_combo"],
            m_pos_station=float(row["m_pos_station"]),
            m_neg=float(row["m_neg"]), m_neg_combo=row["m_neg_combo"],
            m_neg_station=float(row["m_neg_station"]),
            v_max=float(row["shear"]), v_combo="",
        )
        d = StripDesign(env=env, h=float(row["h"]),
                        dia_top=row["dia_top"], spacing_top=row["spa_top"],
                        dia_bot=row["dia_bot"], spacing_bot=row["spa_bot"])
        designs.append(d)
    mat = MaterialParams(mode=CalcMode.EXCEL_COMPAT)
    for d in designs:
        calc_strip(d, mat)

    out = tmp_path / "baocao.xlsx"
    export_report(designs, mat, str(out), source="test.mdb")
    return load_workbook(str(out)), designs


def test_tieu_de_va_thong_tin(report):
    wb, _ = report
    ws = wb.active
    assert ws["B1"].value == "TÍNH TOÁN THÉP ĐÀI CỌC"
    texts = [ws.cell(row=r, column=2).value or "" for r in range(2, 7)]
    joined = " ".join(str(t) for t in texts)
    assert "TCVN 5574" in joined
    assert "B22.5" in joined and "CB400" in joined
    assert "Giống Excel gốc" in joined


def test_bang_du_lieu_khop_golden(report, golden_d1):
    wb, designs = report
    ws = wb.active
    # tìm dòng header (cột B = "STT")
    head = next(r for r in range(1, 20) if ws.cell(row=r, column=2).value == "STT")
    first = head + 1
    g0 = golden_d1["rows"][0]
    assert ws.cell(row=first, column=5).value == g0["strip"]                      # E: Strip
    assert ws.cell(row=first, column=9).value == pytest.approx(g0["m_pos"])       # I: M+
    assert ws.cell(row=first, column=14).value == pytest.approx(float(g0["as_top"]))   # N: Astop
    # đủ số dòng + chân trang có version
    assert ws.cell(row=head + len(designs), column=5).value == golden_d1["rows"][-1]["strip"]
    foot_texts = [str(ws.cell(row=r, column=2).value or "") for r in range(head + len(designs), head + len(designs) + 4)]
    assert any("rebarFlow v" in t for t in foot_texts)


def test_check_to_mau(report):
    wb, designs = report
    ws = wb.active
    head = next(r for r in range(1, 20) if ws.cell(row=r, column=2).value == "STT")
    # dòng nào check_top là "CT" → nền cam FFEB9C; là số >=1 → xanh C6EFCE
    for i, d in enumerate(designs, start=1):
        c = ws.cell(row=head + i, column=20)  # T: check trên
        if d.check_top == "CT":
            assert c.fill.start_color.rgb.endswith("FFEB9C")
        elif isinstance(d.check_top, float) and d.check_top >= 1:
            assert c.fill.start_color.rgb.endswith("C6EFCE")
