"""Trích số liệu kỳ vọng (golden) từ file Excel gốc → tests/golden/*.json.

Chạy 1 lần (hoặc khi file gốc thay đổi):

    python tests/tools/extract_golden.py

Input : tests/fixtures/tinh Dai coc - GOC.xlsx
        (convert từ .xls bằng: soffice --headless --convert-to xlsx "<file.xls>")
Output: tests/golden/golden_d1.json          — 33 dòng kết quả sheet D1 (output tinhthepmax)
        tests/golden/strip_forces_sample.json — sheet Data (input tương ứng)
"""

import json
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[2]
XLSX = ROOT / "tests" / "fixtures" / "tinh Dai coc - GOC.xlsx"
OUT_DIR = ROOT / "tests" / "golden"


def main() -> None:
    wb = openpyxl.load_workbook(XLSX, data_only=True)  # data_only: giá trị đã tính sẵn

    d1 = wb["D1"]
    params = {
        "concrete": d1["I9"].value,          # "B22.5"
        "steel": d1["I10"].value,            # "CB400"
        "cover_top_mm": d1["F12"].value,     # 35
        "cover_bot_mm": d1["F13"].value,     # 100
        "as_ham_cm2": d1["F16"].value or 0,  # 0
    }

    rows = []
    r = 21
    while d1.cell(row=r, column=5).value not in (None, ""):
        rows.append({
            "h": d1[f"D{r}"].value,
            "strip": d1[f"E{r}"].value,
            "width": d1[f"F{r}"].value,
            "m_pos_combo": d1[f"G{r}"].value,
            "m_pos_station": d1[f"H{r}"].value,
            "m_pos": d1[f"I{r}"].value,
            "m_neg_combo": d1[f"J{r}"].value,
            "m_neg_station": d1[f"K{r}"].value,
            "m_neg": d1[f"L{r}"].value,
            "shear": d1[f"M{r}"].value,
            "as_top": d1[f"N{r}"].value,     # số hoặc "-"
            "as_bot": d1[f"O{r}"].value,
            "dia_top": d1[f"P{r}"].value,
            "spa_top": d1[f"Q{r}"].value,
            "dia_bot": d1[f"R{r}"].value,
            "spa_bot": d1[f"S{r}"].value,
            "check_top": d1[f"T{r}"].value,  # số, "CT" hoặc "-"
            "check_bot": d1[f"U{r}"].value,
        })
        r += 1

    data = wb["Data"]
    forces = []
    r = 10
    while data.cell(row=r, column=3).value not in (None, ""):
        forces.append({
            "strip": data[f"C{r}"].value,
            "station": data[f"D{r}"].value,
            "location": data[f"E{r}"].value,
            "output_case": data[f"F{r}"].value,
            "case_type": data[f"G{r}"].value,
            "p": data[f"H{r}"].value,
            "v2": data[f"I{r}"].value,
            "t": data[f"J{r}"].value,
            "m3": data[f"K{r}"].value,
            "global_x": data[f"L{r}"].value,
            "global_y": data[f"M{r}"].value,
            "cut_width": data[f"N{r}"].value,
        })
        r += 1

    geo = wb["Object Geometry - Design Strips"]
    geometry = []
    r = 2
    while geo.cell(row=r, column=1).value not in (None, ""):
        geometry.append({
            "strip": geo[f"A{r}"].value,
            "point": geo[f"B{r}"].value,
            "x": geo[f"C{r}"].value,
            "y": geo[f"D{r}"].value,
        })
        r += 1

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    (OUT_DIR / "geometry_sample.json").write_text(
        json.dumps(geometry, indent=1, ensure_ascii=False), encoding="utf-8"
    )
    (OUT_DIR / "golden_d1.json").write_text(
        json.dumps({"params": params, "rows": rows}, indent=1, ensure_ascii=False),
        encoding="utf-8",
    )
    (OUT_DIR / "strip_forces_sample.json").write_text(
        json.dumps(forces, indent=1, ensure_ascii=False), encoding="utf-8"
    )
    print(f"OK: {len(rows)} dòng golden D1, {len(forces)} dòng strip forces, "
          f"{len(geometry)} dòng geometry")


if __name__ == "__main__":
    main()
